#!/usr/bin/env python

#
# Generate/update a XLSX BOM from a KiCad generic netlist
#

"""
    @package
    Generate/update a XLSX BOM.
    Components are sorted by ref and grouped by value with same footprint
    Fields are (if exist)
    'Ref', 'Qty', 'Value', 'Rating', 'Footprint', 'Description', 'MPN', ...

    Command line:
    python "pathToFile/BOM.py" "%I" "%O"
"""

# Import the KiCad python helper module and XLS tools
import netlist_reader
from translate_fp import translate_fp
import openpyxl
import sys

if not len(sys.argv) >= 3:
    print("Command line:\npython \"pathToFile/BOM.py\" bom-from-KiCad.xml generated-bom-xlsx-file")
    print("Command line (from KiCad):\npython \"pathToFile/BOM.py\" \"%I\" \"%O\"")
    sys.exit()

# Generate an instance of a generic netlist, and load the netlist tree from
# the command line option. If the file doesn't exist, execution will stop
net = netlist_reader.netlist(sys.argv[1])

# Open a file to write to, if the file cannot be opened output to stdout
# instead
xlsfile = sys.argv[2] + '.xlsx'


header_names = ['Ref', 'Footprint', 'Value', 'Rating', 'Qty', 'MPN', 'Farnell', 'Mouser']

# Get all of the components in groups of matching value + footprint

# A custom compare / grouping can be defined (for instance named myEqu):
# add the line:
#    kicad_netlist_reader.comp.__eq__ = myEqu
# in this bom generator script before calling the netliste reader by something like:
#    net = kicad_netlist_reader.netlist(sys.argv[1])
# see netlist_reader.py for more info
grouped = net.groupComponents()


def init_BOM_sheet(xls):
    sheet = xls['BOM']
    sheet.insert_rows(1)
    headers = ['Sync'] + header_names
    for i, col in enumerate(headers):
        col_no = i + 1
        cell = sheet.cell(column=col_no, row=1)
        cell.value = col
        cell.font = cell.font.copy(bold=True)

try:
    xls = openpyxl.load_workbook(xlsfile)

# Create a new empty sheet with only the headers.
# The rest of the sync process will add all the data
except FileNotFoundError:
    xls = openpyxl.Workbook()
    xls.active.title = 'BOM'
    init_BOM_sheet(xls)

if not 'BOM' in xls:
    print("WARNING: xls file {} did not contain a 'BOM' worksheet, adding new sheet..".format(xlsfile))
    xls.create_sheet('BOM')
    init_BOM_sheet(xls)

# Build a lookup from column header -> column index
sheet = xls['BOM']
col_lookup = {}
col = 0
while(True):
    col+=1
    val = sheet.cell(column=col,row=1).value
    if val is None:
        break
    col_lookup[val] = (col - 1)

if not 'Value' in col_lookup or not 'Footprint' in col_lookup:
    print("ERROR: xls file {} does not have 'Value' and 'Footprint' colums".format(xlsfile))
    sys.exit()

## XLS: styling

# New part
new_color = openpyxl.styles.colors.Color(rgb='0000F200')
new_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=new_color)

# Changed part
changed_color = openpyxl.styles.colors.Color(rgb='00FFF200')
changed_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=changed_color)

# Translated content (e.g. simplfied footprint name)
translate_color = openpyxl.styles.colors.Color(rgb='007DF2E6')
translate_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=translate_color)

# Obsolete part: part is not in the Kicad design anymore
obsolete_color = openpyxl.styles.colors.Color(rgb='00F20000')
obsolete_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=obsolete_color)

# Default style
none_fill = openpyxl.styles.fills.PatternFill(patternType=None)

## XLS: prepare by clearing the sync column
if 'Sync' in col_lookup:
    col_no = col_lookup['Sync']+1
    for r in range(sheet.max_row)[1:]:
        row_no = r+1
        cell = sheet.cell(column=col_no, row=row_no)
        cell.value = None
        cell.fill = none_fill
else:
    print("WARNING: xls file {} does not have a 'Sync' column."
            "This means you cannot detect obsolete entries...".format(xlsfile))



last_updated_row = 1
def update_xls(part):
    global last_updated_row

    for row in sheet.iter_rows():

        xls_val = row[col_lookup['Value']].value
        xls_fp = row[col_lookup['Footprint']].value

        # Not the right row: skip
        if not xls_val == part['Value'] or not translate_fp(xls_fp) == translate_fp(part['Footprint']):
            continue

        # Matching row found: mark it as 'in sync'
        if 'Sync' in col_lookup:
            row[col_lookup['Sync']].value = 1

        # Check each property agains the XLS value in the corresponding column
        first_change = True
        for prop in part:
            new_value = str(part[prop]).strip()

            if prop == 'Footprint':
                new_value = translate_fp(new_value)

            if prop == 'Qty':
                new_value = int(new_value)

            # no value is set: skip
            if not new_value:
                continue

            # column not in sheet: skip
            if not prop in col_lookup:
                continue

            col_index = col_lookup[prop]
            old_value = str(row[col_index].value).strip()
            if not old_value == str(new_value):
                if first_change:
                    first_change = False
                    print("Change(s) found for component with value='{}', "
                            "footprint '{}':".format(xls_val, xls_fp))
                print("'{}' changed from '{}' to '{}'".format(prop, old_value, new_value))

                # This can only be because of changes in translation,
                # otherwise this row would not have matched
                if prop == 'Footprint':
                    print("Translated")
                    row[col_index].fill = translate_fill
                else:
                    row[col_index].fill = changed_fill

                row[col_index].value = new_value
        last_updated_row = row[0].row
        return

    # No matching row was found: insert a new one
    print("New component found with value='{}', "
            "footprint '{}':".format(part['Value'], part['Footprint']))

    last_updated_row+=1

    if len(list(sheet.rows)) < last_updated_row:
        # Append new row to sheet by writing a dummy value
        # (dummy value is overwritten, but forces sheet.rows[i] to exist)
        sheet.cell(row=last_updated_row,column=1).value='dummy'
    else:
        # Insert new row before given index
        sheet.insert_rows(last_updated_row)


    for prop in part:
        new_value = str(part[prop]).strip()

        # no value is set: skip
        if not new_value:
            continue

        # column not in sheet: skip
        if not prop in col_lookup:
            continue

        row = list(sheet.rows)[last_updated_row-1]

        # Footprint is 'special': translate it to more readable format
        if prop == 'Footprint':
            new_value = translate_fp(new_value)

        if prop == 'Qty':
            new_value = int(new_value)

        # update property
        col_index = col_lookup[prop]
        row[col_index].fill = new_fill
        row[col_index].value = new_value

        # Matching row found: mark it as 'in sync'
        if 'Sync' in col_lookup:
            row[col_lookup['Sync']].fill = new_fill
            row[col_lookup['Sync']].value = 1



## Output all of the component information
for group in grouped:
    refs = ""

    # Add the reference of every component in the group and keep a reference
    # to the component so that the other data can be filled in once per group

    ratings = set()

    for component in group:
        refs += component.getRef() + ", "
        c = component

        # Gather all component ratings for this group of components.
        # All unique ratings are combined into one field in the BOM.
        # While ordering, a component should be selected that satisfies all of them
        rating = str(c.getField("Rating") or c.getField("rating")).strip()
        for r in rating.split(','):
            if len(r):
                ratings.add(r)

    ratings = ','.join(list(ratings))

    part = {}
    part['Ref'] = refs
    part['Qty'] = len(group)
    part['Value'] = c.getValue()
    part['Footprint'] = c.getFootprint()
    part['Description'] = c.getDescription()
    part['Rating'] = ratings
    part['MPN'] = c.getField("MPN")
    part['Farnell'] = c.getField("Farnell")
    part['Mouser'] = c.getField("Mouser")
    part['DNI'] = c.getField("DNI")

    # Avoid whitespace mismatch
    for prop in part:
        part[prop] = str(part[prop]).strip()

    # Skip 'Do Not Place' parts
    if part['Value'] in ['DNI', 'DNP', 'LOGO', 'mousebite', 'inf']:
        continue
    if part['Value'].startswith('DNI'):
        continue
    if part['DNI']:
        continue



    update_xls(part)



## XLS: style all obsolete entries
if 'Sync' in col_lookup:
    print("Styling obsolete entries")
    col_no = col_lookup['Sync']+1
    for r in range(sheet.max_row)[1:]:
        row_no = r+1
        cell = sheet.cell(column=col_no, row=row_no)
        if not cell.value:
            val = sheet.cell(column=col_lookup['Value']+1, row=row_no).value
            fp = sheet.cell(column=col_lookup['Footprint']+1, row=row_no).value
            if val or fp:
                
                print("Obsolete component found with value='{}', "
                        "footprint '{}':".format(val, fp))

                cell.fill = obsolete_fill

xls.save(filename=xlsfile)

